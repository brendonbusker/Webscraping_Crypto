from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font
import keys
from twilio.rest import Client

# Configure Twilio
client = Client(keys.account_sid, keys.auth_token)

TWnumber = '+17064206923'
myphone=   '+15126604148'

# Configure Webscrape
webpage = 'https://www.webull.com/quote/crypto'
page = urlopen(webpage)			
soup = BeautifulSoup(page, 'html.parser')
title = soup.title
stock_data = soup.findAll('div', class_ = 'table-cell')

# Configure Excel Sheet
wb = xl.Workbook()
ws = wb.active  
ws.title = 'Crypto Data'
title_font = Font(name='Calibri', size=18, italic=False, bold=True)
negative_font = Font(name='Calibri', size = 14, color='FF0000')
positive_font = Font(name='Calibri', size = 14, color='00FF00')
reg_font = Font(name='Calibri', size = 14)
write_sheet = wb['Crypto Data']

ws['A1'] = "Number"
ws['A1'].font = title_font
ws.column_dimensions['A'].width = 15

ws['B1'] = "Name"
ws['B1'].font = title_font
ws.column_dimensions['B'].width = 35

ws['C1'] = "Symbol"
ws['C1'].font = title_font
ws.column_dimensions['C'].width = 15

ws['D1'] = "Current Price"
ws['D1'].font = title_font
ws.column_dimensions['D'].width = 25

ws['E1'] = "Percent Change"
ws['E1'].font = title_font
ws.column_dimensions['E'].width = 25

ws['F1'] = "Corresponding Price"
ws['F1'].font = title_font
ws.column_dimensions['F'].width = 30

# Begins loop through website and initializes counters
stock_counter = 1
excel_counter = 1
price_list = []

for x in range(38):
    # This splits up the Symbol from the name
    name = stock_data[stock_counter].text
    symbol = name[1:name.index('USD')+3]

    # Remove symbol from name and repeated first character
    name = name.replace(symbol, '')[1:]

    # This catches errors when a stock has no data
    try:
        current_price = float(stock_data[stock_counter+1].text.replace(',', ''))

    except: #takes into account some crypto is filled with '--'
        current_price = stock_data[stock_counter+1].text.replace(',', '')

    percent_change = stock_data[stock_counter+2].text

    # This catches errors when a stock has no data
    try:
        change = round((float(percent_change.replace('%', '')) / 100) * current_price, 4)
        
    except:
        pass
    
    # Sends text message if btc or eth inc or dec within 5 dollars
    if symbol == 'BTCUSD':
        if change >= -5 and change <= 5:
            textmsg = client.messages.create(to=myphone, from_=TWnumber, body='BTC has increased/decreased within $5')
    if symbol == 'ETHUSD':
        if change >= -5 and change <= 5:
            textmsg = client.messages.create(to=myphone, from_=TWnumber, body='ETH has increased/decreased within $5')

    # Ensures it only displays the top 5 on WeBull   
    if stock_counter < 51:
        # Add to excel
        ws['A' + str(excel_counter + 1)] = excel_counter
        ws['A' + str(excel_counter + 1)].font = reg_font
        ws['B' + str(excel_counter + 1)] = name
        ws['B' + str(excel_counter + 1)].font = reg_font
        ws['C' + str(excel_counter + 1)] = symbol
        ws['C' + str(excel_counter + 1)].font = reg_font


        if '-' in percent_change:
            ws['D' + str(excel_counter + 1)] = '$' + str(round(current_price, 4))
            ws['D' + str(excel_counter + 1)].font = negative_font
            ws['E' + str(excel_counter + 1)] = percent_change
            ws['E' + str(excel_counter + 1)].font = negative_font
            ws['F' + str(excel_counter + 1)] = '$' + str(round(change + current_price, 4))
            ws['F' + str(excel_counter + 1)].font = negative_font

        else:
            ws['D' + str(excel_counter + 1)] = '$' + str(round(current_price, 4))
            ws['D' + str(excel_counter + 1)].font = positive_font
            ws['E' + str(excel_counter + 1)] = percent_change
            ws['E' + str(excel_counter + 1)].font = positive_font
            ws['F' + str(excel_counter + 1)] = '$' + str(round(change + current_price, 4))
            ws['F' + str(excel_counter + 1)].font = positive_font
        
    excel_counter += 1
    stock_counter += 10

print('Success! Check folder for CryptoData.xlsx')
wb.save('CryptoData.xlsx')


